from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib import messages
from django.db.models import Q, Count, Avg, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Trade, WeeklyReview, MonthlyReview, TradingPsychology, TradingGoal, MarketCondition, TradingHabit
from .forms import TradeForm, WeeklyReviewForm, MonthlyReviewForm, TradeFilterForm, CustomUserCreationForm, TradingPsychologyForm, TradingGoalForm, MarketConditionForm, TradingHabitForm
from .supabase_client import upload_file_to_supabase, delete_file_from_supabase
import json
import csv
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.units import inch


def register_view(request):
    """User registration view"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful! Welcome to your Trading Journal.')
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})




@login_required
def dashboard(request):
    """Main dashboard with analytics and performance metrics"""
    user_trades = Trade.objects.filter(user=request.user)
    
    # Basic statistics
    total_trades = user_trades.count()
    winning_trades = user_trades.filter(profit_loss__gt=0).count()
    losing_trades = user_trades.filter(profit_loss__lt=0).count()
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    
    # P&L statistics
    total_pnl = user_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
    avg_profit = user_trades.filter(profit_loss__gt=0).aggregate(avg=Avg('profit_loss'))['avg'] or 0
    avg_loss = user_trades.filter(profit_loss__lt=0).aggregate(avg=Avg('profit_loss'))['avg'] or 0
    
    # Recent trades
    recent_trades = user_trades[:5]
    
    # Monthly performance data for charts
    monthly_data = []
    for i in range(6):  # Last 6 months
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start.replace(day=28) + timedelta(days=4)
        month_end = month_end - timedelta(days=month_end.day)
        
        month_trades = user_trades.filter(date__gte=month_start, date__lte=month_end)
        month_pnl = month_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'pnl': month_pnl,
            'trades': month_trades.count()
        })
    
    monthly_data.reverse()
    
    # Setup performance
    setup_performance = user_trades.values('setup_type').annotate(
        count=Count('id'),
        total_pnl=Sum('profit_loss'),
        avg_pnl=Avg('profit_loss')
    ).order_by('-total_pnl')
    
    # Risk-reward analysis
    trades_with_rr = user_trades.exclude(stop_loss=0).exclude(target_price=0)
    avg_risk_reward = 0
    if trades_with_rr.exists():
        rr_ratios = []
        for trade in trades_with_rr:
            rr = trade.get_risk_reward_ratio()
            if rr:
                rr_ratios.append(rr)
        avg_risk_reward = sum(rr_ratios) / len(rr_ratios) if rr_ratios else 0
    
    context = {
        'total_trades': total_trades,
        'winning_trades': winning_trades,
        'losing_trades': losing_trades,
        'win_rate': round(win_rate, 1),
        'total_pnl': round(total_pnl, 2),
        'avg_profit': round(avg_profit, 2),
        'avg_loss': round(avg_loss, 2),
        'avg_risk_reward': round(avg_risk_reward, 2),
        'recent_trades': recent_trades,
        'monthly_data': json.dumps(monthly_data),
        'setup_performance': setup_performance,
    }
    
    return render(request, 'journal/dashboard.html', context)


@login_required
def trade_list(request):
    """List all trades with filtering and pagination"""
    trades = Trade.objects.filter(user=request.user)
    
    # Apply filters
    filter_form = TradeFilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['symbol']:
            trades = trades.filter(symbol__icontains=filter_form.cleaned_data['symbol'])
        if filter_form.cleaned_data['trade_type']:
            trades = trades.filter(trade_type=filter_form.cleaned_data['trade_type'])
        if filter_form.cleaned_data['setup_type']:
            trades = trades.filter(setup_type=filter_form.cleaned_data['setup_type'])
        if filter_form.cleaned_data['date_from']:
            trades = trades.filter(date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data['date_to']:
            trades = trades.filter(date__lte=filter_form.cleaned_data['date_to'])
        if filter_form.cleaned_data['profitable_only']:
            trades = trades.filter(profit_loss__gt=0)
        if filter_form.cleaned_data['loss_only']:
            trades = trades.filter(profit_loss__lt=0)
    
    # Calculate total P&L for filtered trades
    total_pnl = trades.aggregate(total=Sum('profit_loss'))['total'] or 0
    total_trades = trades.count()
    winning_trades = trades.filter(profit_loss__gt=0).count()
    losing_trades = trades.filter(profit_loss__lt=0).count()
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    
    # Pagination
    paginator = Paginator(trades, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filter_form': filter_form,
        'total_pnl': total_pnl,
        'total_trades': total_trades,
        'winning_trades': winning_trades,
        'losing_trades': losing_trades,
        'win_rate': win_rate,
    }
    
    return render(request, 'journal/trade_list.html', context)


@login_required
def trade_detail(request, pk):
    """View individual trade details"""
    trade = get_object_or_404(Trade, pk=pk, user=request.user)
    return render(request, 'journal/trade_detail.html', {'trade': trade})


@login_required
def trade_create(request):
    """Create a new trade"""
    if request.method == 'POST':
        form = TradeForm(request.POST, request.FILES)
        if form.is_valid():
            trade = form.save(commit=False)
            trade.user = request.user
            
            # Handle file uploads to Supabase
            if 'screenshot_before' in request.FILES:
                file = request.FILES['screenshot_before']
                file_path = f"before/{request.user.id}/{trade.symbol}_{trade.date}_{file.name}"
                file_content = file.read()
                
                # Upload to Supabase
                supabase_url = upload_file_to_supabase(
                    file_content, 
                    file_path, 
                    file.content_type
                )
                if supabase_url:
                    trade.screenshot_before_url = supabase_url
            
            if 'screenshot_after' in request.FILES:
                file = request.FILES['screenshot_after']
                file_path = f"after/{request.user.id}/{trade.symbol}_{trade.date}_{file.name}"
                file_content = file.read()
                
                # Upload to Supabase
                supabase_url = upload_file_to_supabase(
                    file_content, 
                    file_path, 
                    file.content_type
                )
                if supabase_url:
                    trade.screenshot_after_url = supabase_url
            
            trade.save()
            messages.success(request, 'Trade added successfully!')
            return redirect('trade_detail', pk=trade.pk)
    else:
        form = TradeForm()
    
    return render(request, 'journal/trade_form.html', {
        'form': form,
        'title': 'Add New Trade'
    })


@login_required
def trade_edit(request, pk):
    """Edit an existing trade"""
    trade = get_object_or_404(Trade, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TradeForm(request.POST, request.FILES, instance=trade)
        if form.is_valid():
            form.save()
            messages.success(request, 'Trade updated successfully!')
            return redirect('trade_detail', pk=trade.pk)
    else:
        form = TradeForm(instance=trade)
    
    return render(request, 'journal/trade_form.html', {
        'form': form,
        'trade': trade,
        'title': 'Edit Trade'
    })


@login_required
def trade_delete(request, pk):
    """Delete a trade"""
    trade = get_object_or_404(Trade, pk=pk, user=request.user)
    
    if request.method == 'POST':
        trade.delete()
        messages.success(request, 'Trade deleted successfully!')
        return redirect('trade_list')
    
    return render(request, 'journal/trade_confirm_delete.html', {'trade': trade})


@login_required
def weekly_reviews(request):
    """List weekly reviews"""
    reviews = WeeklyReview.objects.filter(user=request.user)
    return render(request, 'journal/weekly_reviews.html', {'reviews': reviews})


@login_required
def weekly_review_create(request):
    """Create a new weekly review"""
    if request.method == 'POST':
        form = WeeklyReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            
            # Calculate weekly statistics
            week_trades = Trade.objects.filter(
                user=request.user,
                date__gte=review.week_start_date,
                date__lte=review.week_end_date
            )
            review.total_trades = week_trades.count()
            review.winning_trades = week_trades.filter(profit_loss__gt=0).count()
            review.losing_trades = week_trades.filter(profit_loss__lt=0).count()
            review.total_pnl = week_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
            
            review.save()
            messages.success(request, 'Weekly review created successfully!')
            return redirect('weekly_reviews')
    else:
        form = WeeklyReviewForm()
    
    return render(request, 'journal/weekly_review_form.html', {
        'form': form,
        'title': 'Create Weekly Review'
    })


@login_required
def monthly_reviews(request):
    """List monthly reviews"""
    reviews = MonthlyReview.objects.filter(user=request.user)
    return render(request, 'journal/monthly_reviews.html', {'reviews': reviews})


@login_required
def monthly_review_create(request):
    """Create a new monthly review"""
    if request.method == 'POST':
        form = MonthlyReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.user = request.user
            
            # Calculate monthly statistics
            month_start = review.month.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
            
            month_trades = Trade.objects.filter(
                user=request.user,
                date__gte=month_start,
                date__lte=month_end
            )
            review.total_trades = month_trades.count()
            review.winning_trades = month_trades.filter(profit_loss__gt=0).count()
            review.losing_trades = month_trades.filter(profit_loss__lt=0).count()
            review.total_pnl = month_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
            
            # Calculate max drawdown (simplified)
            running_pnl = 0
            max_pnl = 0
            max_drawdown = 0
            
            for trade in month_trades.order_by('date'):
                running_pnl += trade.profit_loss
                max_pnl = max(max_pnl, running_pnl)
                drawdown = max_pnl - running_pnl
                max_drawdown = max(max_drawdown, drawdown)
            
            review.max_drawdown = max_drawdown
            review.save()
            messages.success(request, 'Monthly review created successfully!')
            return redirect('monthly_reviews')
    else:
        form = MonthlyReviewForm()
    
    return render(request, 'journal/monthly_review_form.html', {
        'form': form,
        'title': 'Create Monthly Review'
    })


@login_required
def analytics(request):
    """Advanced analytics page"""
    user_trades = Trade.objects.filter(user=request.user)
    
    # Performance by setup type
    setup_stats = user_trades.values('setup_type').annotate(
        count=Count('id'),
        total_pnl=Sum('profit_loss'),
        avg_pnl=Avg('profit_loss'),
        win_rate=Count('id', filter=Q(profit_loss__gt=0)) * 100.0 / Count('id')
    ).order_by('-total_pnl')
    
    # Performance by symbol
    symbol_stats = user_trades.values('symbol').annotate(
        count=Count('id'),
        total_pnl=Sum('profit_loss'),
        avg_pnl=Avg('profit_loss')
    ).order_by('-total_pnl')[:10]
    
    # Daily performance
    daily_stats = user_trades.values('date').annotate(
        count=Count('id'),
        total_pnl=Sum('profit_loss')
    ).order_by('-date')[:30]
    
    # Confidence level analysis
    confidence_stats = user_trades.values('confidence_level').annotate(
        count=Count('id'),
        total_pnl=Sum('profit_loss'),
        avg_pnl=Avg('profit_loss')
    ).order_by('confidence_level')
    
    context = {
        'setup_stats': setup_stats,
        'symbol_stats': symbol_stats,
        'daily_stats': daily_stats,
        'confidence_stats': confidence_stats,
    }
    
    return render(request, 'journal/analytics.html', context)


@login_required
def export_trades_csv(request):
    """Export trades to CSV"""
    user_trades = Trade.objects.filter(user=request.user).order_by('-date')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="trading_journal_{request.user.username}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Symbol', 'Type', 'Entry Price', 'Exit Price', 'Quantity', 'P&L', 'Percentage', 'Setup', 'Confidence', 'Exit Reason'])
    
    for trade in user_trades:
        writer.writerow([
            trade.date.strftime('%Y-%m-%d'),
            trade.symbol,
            trade.trade_type,
            trade.entry_price,
            trade.exit_price,
            trade.quantity,
            trade.profit_loss,
            trade.percentage_gain_loss,
            trade.setup_type,
            trade.confidence_level,
            trade.exit_reason
        ])
    
    return response


@login_required
def export_trades_pdf(request):
    """Export trades to PDF with advanced professional styling and charts"""
    user_trades = Trade.objects.filter(user=request.user).order_by('-date')
    
    # Calculate comprehensive stats
    total_trades = user_trades.count()
    winning_trades = user_trades.filter(profit_loss__gt=0).count()
    losing_trades = user_trades.filter(profit_loss__lt=0).count()
    total_pnl = user_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    avg_profit = user_trades.filter(profit_loss__gt=0).aggregate(avg=Avg('profit_loss'))['avg'] or 0
    avg_loss = user_trades.filter(profit_loss__lt=0).aggregate(avg=Avg('profit_loss'))['avg'] or 0
    
    # Create PDF response with proper headers for browser compatibility
    response = HttpResponse(content_type='application/pdf')
    
    # Check if download parameter is present
    if request.GET.get('download'):
        response['Content-Disposition'] = f'attachment; filename="trading_report_{request.user.username}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    else:
        response['Content-Disposition'] = f'inline; filename="trading_report_{request.user.username}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    response['X-Content-Type-Options'] = 'nosniff'
    
    # Create PDF with custom page size
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=50, leftMargin=50, 
                           topMargin=80, bottomMargin=50)
    styles = getSampleStyleSheet()
    story = []
    
    # Advanced custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=32,
        spaceAfter=25,
        alignment=1,
        textColor=colors.HexColor('#1E40AF'),
        fontName='Helvetica-Bold',
        borderWidth=2,
        borderColor=colors.HexColor('#3B82F6'),
        borderPadding=10,
        backColor=colors.HexColor('#EFF6FF')
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=18,
        spaceAfter=12,
        alignment=1,
        textColor=colors.HexColor('#059669'),
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#ECFDF5'),
        borderWidth=1,
        borderColor=colors.HexColor('#10B981'),
        borderPadding=8
    )
    
    # Create professional header with gradient effect
    header_data = [
        ['📊 TRADING JOURNAL REPORT'],
        [f'👤 Trader: {request.user.username}'],
        [f'📅 Generated: {timezone.now().strftime("%B %d, %Y at %I:%M %p")}'],
        [f'📈 Total Trades: {total_trades} | 💰 Total P&L: ₹{total_pnl:.2f}']
    ]
    
    header_table = Table(header_data, colWidths=[500])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EFF6FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1E40AF')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 24),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 20),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#3B82F6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 30))
    
    # Performance Overview with enhanced styling
    story.append(Paragraph("🎯 PERFORMANCE OVERVIEW", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Enhanced metrics with better visual design
    metrics_data = [
        ['📈 TOTAL TRADES', f'{total_trades}', '📊 WIN RATE', f'{win_rate:.1f}%'],
        ['🏆 WINNING TRADES', f'{winning_trades}', '💔 LOSING TRADES', f'{losing_trades}'],
        ['💰 TOTAL P&L', f'₹{total_pnl:.2f}', '📊 AVG PROFIT', f'₹{avg_profit:.2f}'],
        ['📉 AVG LOSS', f'₹{avg_loss:.2f}', '⚖️ RISK/REWARD', f'{abs(avg_profit/avg_loss):.2f}:1' if avg_loss != 0 else 'N/A']
    ]
    
    metrics_table = Table(metrics_data, colWidths=[140, 100, 140, 100])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#3B82F6')),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#EFF6FF')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#10B981')),
        ('BACKGROUND', (3, 0), (3, -1), colors.HexColor('#ECFDF5')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.white),
        ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#059669')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 2, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
    ]))
    
    story.append(metrics_table)
    story.append(Spacer(1, 30))
    
    # Performance Analysis with enhanced insights
    if total_trades > 0:
        story.append(Paragraph("📊 PERFORMANCE ANALYSIS", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Enhanced insights with better formatting
        insights_data = []
        
        # Win rate analysis
        if win_rate >= 70:
            insights_data.append(['🎉 EXCELLENT PERFORMANCE', f'Win Rate: {win_rate:.1f}% - You are trading like a professional!'])
        elif win_rate >= 50:
            insights_data.append(['👍 GOOD PERFORMANCE', f'Win Rate: {win_rate:.1f}% - Keep up the consistent performance!'])
        else:
            insights_data.append(['💪 IMPROVEMENT OPPORTUNITY', f'Win Rate: {win_rate:.1f}% - Focus on better setups and risk management!'])
        
        # P&L analysis
        if total_pnl > 0:
            insights_data.append(['💰 PROFITABLE TRADER', f'Total P&L: ₹{total_pnl:.2f} - Your strategy is working excellently!'])
        else:
            insights_data.append(['📈 LEARNING PHASE', f'Total P&L: ₹{total_pnl:.2f} - Every loss is a valuable lesson!'])
        
        # Risk management analysis
        if avg_profit > abs(avg_loss) and avg_loss != 0:
            insights_data.append(['⚖️ EXCELLENT RISK MANAGEMENT', f'Risk/Reward: {abs(avg_profit/avg_loss):.2f}:1 - Your wins are bigger than losses!'])
        else:
            insights_data.append(['🎯 RISK MANAGEMENT FOCUS', 'Work on cutting losses quickly and letting profits run!'])
        
        # Confidence analysis
        avg_confidence = user_trades.aggregate(avg=Avg('confidence_level'))['avg'] or 0
        if avg_confidence >= 8:
            insights_data.append(['⭐ HIGH CONFIDENCE TRADER', f'Average Confidence: {avg_confidence:.1f}/10 - You trust your analysis!'])
        elif avg_confidence >= 6:
            insights_data.append(['📊 MODERATE CONFIDENCE', f'Average Confidence: {avg_confidence:.1f}/10 - Good balance of caution and confidence!'])
        else:
            insights_data.append(['🤔 LOW CONFIDENCE', f'Average Confidence: {avg_confidence:.1f}/10 - Work on building confidence through better analysis!'])
        
        insights_table = Table(insights_data, colWidths=[200, 300])
        insights_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F59E0B')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FEF3C7')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#92400E')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#F59E0B')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(insights_table)
        story.append(Spacer(1, 30))
    
    # Page break for detailed trades
    story.append(PageBreak())
    
    # Detailed Trades Section with enhanced styling
    if user_trades.exists():
        story.append(Paragraph("📋 DETAILED TRADE HISTORY", subtitle_style))
        story.append(Spacer(1, 20))
        
        trades_data = [['📅 DATE', '🏷️ SYMBOL', '📊 TYPE', '⬆️ ENTRY', '⬇️ EXIT', '💰 P&L', '🎯 SETUP', '⭐ CONF']]
        
        for trade in user_trades[:25]:  # Show last 25 trades
            trades_data.append([
                trade.date.strftime('%d/%m/%Y'),
                trade.symbol,
                trade.trade_type,
                f"₹{trade.entry_price:.2f}",
                f"₹{trade.exit_price:.2f}",
                f"₹{trade.profit_loss:.2f}",
                trade.setup_type.title(),
                f"{trade.confidence_level}/10"
            ])
        
        trades_table = Table(trades_data, colWidths=[70, 60, 50, 70, 70, 80, 80, 50])
        trades_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#3B82F6')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
        ]))
        
        story.append(trades_table)
        story.append(Spacer(1, 30))
    
    # Setup Performance Analysis with enhanced design
    if user_trades.exists():
        story.append(Paragraph("🎯 SETUP PERFORMANCE ANALYSIS", subtitle_style))
        story.append(Spacer(1, 20))
        
        setup_stats = user_trades.values('setup_type').annotate(
            count=Count('id'),
            total_pnl=Sum('profit_loss'),
            avg_pnl=Avg('profit_loss')
        ).order_by('-total_pnl')
        
        if setup_stats:
            setup_data = [['🎯 SETUP TYPE', '📊 TRADES', '💰 TOTAL P&L', '📈 AVG P&L', '📊 SUCCESS RATE']]
            
            for setup in setup_stats[:8]:
                setup_trades = user_trades.filter(setup_type=setup['setup_type'])
                setup_wins = setup_trades.filter(profit_loss__gt=0).count()
                setup_success_rate = (setup_wins / setup['count'] * 100) if setup['count'] > 0 else 0
                
                setup_data.append([
                    setup['setup_type'].title(),
                    str(setup['count']),
                    f"₹{setup['total_pnl']:.2f}",
                    f"₹{setup['avg_pnl']:.2f}",
                    f"{setup_success_rate:.1f}%"
                ])
            
            setup_table = Table(setup_data, colWidths=[100, 60, 90, 90, 80])
            setup_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#10B981')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDF4')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
                ('TOPPADDING', (0, 1), (-1, -1), 12),
            ]))
            
            story.append(setup_table)
            story.append(Spacer(1, 30))
    
    # Monthly Performance Summary
    if user_trades.exists():
        story.append(Paragraph("📅 MONTHLY PERFORMANCE", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Calculate monthly performance
        monthly_data = []
        current_month = timezone.now().replace(day=1)
        
        for i in range(6):  # Last 6 months
            month_start = current_month - timedelta(days=30*i)
            month_end = month_start.replace(day=28) + timedelta(days=4)
            month_end = month_end - timedelta(days=month_end.day)
            
            month_trades = user_trades.filter(date__gte=month_start, date__lte=month_end)
            month_pnl = month_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
            month_count = month_trades.count()
            
            if month_count > 0:
                monthly_data.append([
                    month_start.strftime('%B %Y'),
                    str(month_count),
                    f"₹{month_pnl:.2f}",
                    f"{(month_trades.filter(profit_loss__gt=0).count() / month_count * 100):.1f}%"
                ])
        
        if monthly_data:
            monthly_data.insert(0, ['📅 MONTH', '📊 TRADES', '💰 P&L', '📈 WIN RATE'])
            
            monthly_table = Table(monthly_data, colWidths=[120, 80, 100, 80])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#8B5CF6')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAF5FF')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
                ('TOPPADDING', (0, 1), (-1, -1), 12),
            ]))
            
            story.append(monthly_table)
            story.append(Spacer(1, 30))
    
    # Professional Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=12,
        spaceBefore=30,
        alignment=1,
        textColor=colors.HexColor('#6B7280'),
        fontName='Helvetica-Oblique',
        backColor=colors.HexColor('#F9FAFB'),
        borderWidth=1,
        borderColor=colors.HexColor('#E5E7EB'),
        borderPadding=15
    )
    
    story.append(Paragraph("📊 Generated by Trading Journal - Your Professional Trading Companion", footer_style))
    story.append(Paragraph(f"🕒 Report generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", footer_style))
    story.append(Paragraph("💡 Keep trading, keep learning, keep growing!", footer_style))
    
    try:
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        
        # Ensure PDF has content
        if len(pdf) < 1000:  # PDF should be at least 1KB
            error_response = HttpResponse("PDF generation failed: Generated PDF is too small", content_type='text/plain')
            error_response.status_code = 500
            return error_response
            
        response.write(pdf)
        return response
    except Exception as e:
        buffer.close()
        # Return error response
        error_response = HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain')
        error_response.status_code = 500
        return error_response



@login_required
def export_trades_excel(request):
    """Export trades to Excel with professional styling"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    user_trades = Trade.objects.filter(user=request.user).order_by('-date')
    
    # Calculate summary stats
    total_trades = user_trades.count()
    winning_trades = user_trades.filter(profit_loss__gt=0).count()
    losing_trades = user_trades.filter(profit_loss__lt=0).count()
    total_pnl = user_trades.aggregate(total=Sum('profit_loss'))['total'] or 0
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trading Journal"
    
    # Title and header
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value="📊 TRADING JOURNAL REPORT")
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # User info
    ws.merge_cells('A2:K2')
    user_cell = ws.cell(row=2, column=1, value=f"👤 Trader: {request.user.username} | 📅 Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}")
    user_cell.font = Font(bold=True, size=12, color="366092")
    user_cell.alignment = Alignment(horizontal="center")
    
    # Summary section
    ws.merge_cells('A4:K4')
    summary_title = ws.cell(row=4, column=1, value="🎯 PERFORMANCE SUMMARY")
    summary_title.font = Font(bold=True, size=14, color="FFFFFF")
    summary_title.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    summary_title.alignment = Alignment(horizontal="center")
    
    # Summary data
    summary_data = [
        ['📈 Total Trades', total_trades, '📊 Win Rate', f'{win_rate:.1f}%'],
        ['🏆 Winning Trades', winning_trades, '💔 Losing Trades', losing_trades],
        ['💰 Total P&L', f'₹{total_pnl:.2f}', '📊 Avg P&L', f'₹{total_pnl/total_trades:.2f}' if total_trades > 0 else '₹0.00']
    ]
    
    for row_idx, row_data in enumerate(summary_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx % 2 == 1:  # Label columns
                cell.font = Font(bold=True, color="366092")
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            else:  # Value columns
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Trades section header
    trades_start_row = 9
    ws.merge_cells(f'A{trades_start_row}:K{trades_start_row}')
    trades_title = ws.cell(row=trades_start_row, column=1, value="📋 DETAILED TRADE HISTORY")
    trades_title.font = Font(bold=True, size=14, color="FFFFFF")
    trades_title.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    trades_title.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['📅 Date', '🏷️ Symbol', '📊 Type', '⬆️ Entry', '⬇️ Exit', '📊 Qty', '💰 P&L', '📈 %', '🎯 Setup', '⭐ Conf', '📝 Exit Reason']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=trades_start_row+1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Data with enhanced styling
    for row, trade in enumerate(user_trades, trades_start_row+2):
        ws.cell(row=row, column=1, value=trade.date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=trade.symbol)
        ws.cell(row=row, column=3, value=trade.trade_type)
        ws.cell(row=row, column=4, value=trade.entry_price)
        ws.cell(row=row, column=5, value=trade.exit_price)
        ws.cell(row=row, column=6, value=trade.quantity)
        ws.cell(row=row, column=7, value=trade.profit_loss)
        ws.cell(row=row, column=8, value=f"{trade.percentage_gain_loss:.2f}%")
        ws.cell(row=row, column=9, value=trade.setup_type.title())
        ws.cell(row=row, column=10, value=f"{trade.confidence_level}/10")
        ws.cell(row=row, column=11, value=trade.exit_reason)
        
        # Apply styling to each cell
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code based on content
            if col == 3:  # Trade type column
                if trade.trade_type == 'LONG':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            elif col == 7:  # P&L column
                if trade.profit_loss > 0:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True, color="006400")
                elif trade.profit_loss < 0:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    cell.font = Font(bold=True, color="8B0000")
            elif col == 10:  # Confidence column
                if trade.confidence_level >= 8:
                    cell.fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                elif trade.confidence_level >= 6:
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary formulas
    if user_trades.exists():
        last_row = trades_start_row + 1 + user_trades.count()
        
        # Add totals row
        ws.cell(row=last_row+2, column=6, value="TOTAL P&L:").font = Font(bold=True, size=12)
        ws.cell(row=last_row+2, column=7, value=f"₹{total_pnl:.2f}").font = Font(bold=True, size=12, color="006400" if total_pnl > 0 else "8B0000")
        
        ws.cell(row=last_row+3, column=6, value="WIN RATE:").font = Font(bold=True, size=12)
        ws.cell(row=last_row+3, column=7, value=f"{win_rate:.1f}%").font = Font(bold=True, size=12)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="trading_journal_{request.user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response




# Psychology Views
@login_required
def psychology_dashboard(request):
    """Trading psychology dashboard"""
    psychology_records = TradingPsychology.objects.filter(user=request.user).order_by('-date')[:10]
    
    # Calculate psychology stats
    if psychology_records:
        avg_confidence = psychology_records.aggregate(avg=Avg('pre_trade_confidence'))['avg'] or 0
        avg_stress = psychology_records.aggregate(avg=Avg('stress_level'))['avg'] or 0
        avg_sleep = psychology_records.aggregate(avg=Avg('sleep_quality'))['avg'] or 0
        avg_focus = psychology_records.aggregate(avg=Avg('focus_level'))['avg'] or 0
    else:
        avg_confidence = avg_stress = avg_sleep = avg_focus = 0
    
    return render(request, 'journal/psychology_dashboard.html', {
        'psychology_records': psychology_records,
        'avg_confidence': avg_confidence,
        'avg_stress': avg_stress,
        'avg_sleep': avg_sleep,
        'avg_focus': avg_focus,
    })


@login_required
def psychology_create(request):
    """Create psychology record"""
    if request.method == 'POST':
        form = TradingPsychologyForm(request.POST)
        if form.is_valid():
            psychology = form.save(commit=False)
            psychology.user = request.user
            psychology.save()
            messages.success(request, 'Psychology record created successfully!')
            return redirect('psychology_dashboard')
    else:
        form = TradingPsychologyForm()
    
    return render(request, 'journal/psychology_form.html', {'form': form})


@login_required
def psychology_edit(request, pk):
    """Edit psychology record"""
    psychology = get_object_or_404(TradingPsychology, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TradingPsychologyForm(request.POST, instance=psychology)
        if form.is_valid():
            form.save()
            messages.success(request, 'Psychology record updated successfully!')
            return redirect('psychology_dashboard')
    else:
        form = TradingPsychologyForm(instance=psychology)
    
    return render(request, 'journal/psychology_form.html', {'form': form, 'psychology': psychology})


# Goals Views
@login_required
def goals_dashboard(request):
    """Trading goals dashboard"""
    active_goals = TradingGoal.objects.filter(user=request.user, is_active=True).order_by('-created_at')
    achieved_goals = TradingGoal.objects.filter(user=request.user, is_achieved=True).order_by('-updated_at')[:5]
    overdue_goals = TradingGoal.objects.filter(user=request.user, is_active=True, end_date__lt=timezone.now().date(), is_achieved=False)
    
    return render(request, 'journal/goals_dashboard.html', {
        'active_goals': active_goals,
        'achieved_goals': achieved_goals,
        'overdue_goals': overdue_goals,
    })


@login_required
def goal_create(request):
    """Create trading goal"""
    if request.method == 'POST':
        form = TradingGoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user
            goal.save()
            messages.success(request, 'Trading goal created successfully!')
            return redirect('goals_dashboard')
    else:
        form = TradingGoalForm()
    
    return render(request, 'journal/goal_form.html', {'form': form})


@login_required
def goal_edit(request, pk):
    """Edit trading goal"""
    goal = get_object_or_404(TradingGoal, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TradingGoalForm(request.POST, instance=goal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Trading goal updated successfully!')
            return redirect('goals_dashboard')
    else:
        form = TradingGoalForm(instance=goal)
    
    return render(request, 'journal/goal_form.html', {'form': form, 'goal': goal})


# Market Conditions Views
@login_required
def market_conditions(request):
    """Market conditions tracking"""
    conditions = MarketCondition.objects.filter(user=request.user).order_by('-date')[:20]
    
    return render(request, 'journal/market_conditions.html', {
        'conditions': conditions,
    })


@login_required
def market_condition_create(request):
    """Create market condition record"""
    if request.method == 'POST':
        form = MarketConditionForm(request.POST)
        if form.is_valid():
            condition = form.save(commit=False)
            condition.user = request.user
            condition.save()
            messages.success(request, 'Market condition recorded successfully!')
            return redirect('market_conditions')
    else:
        form = MarketConditionForm()
    
    return render(request, 'journal/market_condition_form.html', {'form': form})


# Habits Views
@login_required
def habits_dashboard(request):
    """Trading habits dashboard"""
    good_habits = TradingHabit.objects.filter(user=request.user, habit_type='GOOD', is_active=True)
    bad_habits = TradingHabit.objects.filter(user=request.user, habit_type='BAD', is_active=True)
    
    return render(request, 'journal/habits_dashboard.html', {
        'good_habits': good_habits,
        'bad_habits': bad_habits,
    })


@login_required
def habit_create(request):
    """Create trading habit"""
    if request.method == 'POST':
        form = TradingHabitForm(request.POST)
        if form.is_valid():
            habit = form.save(commit=False)
            habit.user = request.user
            habit.save()
            messages.success(request, 'Trading habit created successfully!')
            return redirect('habits_dashboard')
    else:
        form = TradingHabitForm()
    
    return render(request, 'journal/habit_form.html', {'form': form})


@login_required
def habit_edit(request, pk):
    """Edit trading habit"""
    habit = get_object_or_404(TradingHabit, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TradingHabitForm(request.POST, instance=habit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Trading habit updated successfully!')
            return redirect('habits_dashboard')
    else:
        form = TradingHabitForm(instance=habit)
    
    return render(request, 'journal/habit_form.html', {'form': form, 'habit': habit})
